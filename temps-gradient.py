"""
Calculate and export gradient statistics to Excel with formatting
Includes: percentages, medians, quartiles, percentiles, and patient rankings
"""

import pandas as pd
import numpy as np
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import ColorScaleRule

# Input paths
INVASIVE_INPUT_DIR = Path(r"C:\Users\katia\Desktop\min data\work\data\Patients-infinity\filtered\operation-csv-invasive\aftersai")
NONINVASIVE_INPUT_DIR = Path(r"C:\Users\katia\Desktop\min data\work\data\Patients-infinity\filtered\operation-csv-none-invasive")
OUTPUT_DIR = Path(r"C:\Users\katia\Desktop\min data\work\data\Patients-infinity\filtered\gradient")

OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

# Column names
INV_M_COL = 'ART M (mm(hg)^^ISO+)'
NONINV_M_COL = 'NBP M (mm(hg)^^ISO+)'

GRADIENT_THRESHOLD = 10  # mmHg
TIME_TOLERANCE = 1  # minute


def time_to_minutes(time_str):
    """Convert HH:MM:SS to minutes since start of day"""
    try:
        h, m, s = map(int, str(time_str).split(':'))
        return h * 60 + m + s / 60
    except:
        return np.nan


def filter_noninvasive_by_time_and_value(df_noninv, patient_name):
    """Remove specific non-invasive values based on time ranges and thresholds"""
    if df_noninv is None or len(df_noninv) == 0:
        return df_noninv
    
    df = df_noninv.copy()
    
    if "Patient 29" in patient_name or "Patient_29" in patient_name:
        mask_time = (df['time'] >= '08:45:00') & (df['time'] <= '08:50:00')
        mask_value = (df[NONINV_M_COL] < 45)
        df.loc[mask_time & mask_value, NONINV_M_COL] = np.nan
    
    if "Patient 37" in patient_name or "Patient_37" in patient_name:
        mask1 = (df['time'] >= '11:37:00') & (df['time'] <= '11:52:00') & (df[NONINV_M_COL] < 70)
        df.loc[mask1, NONINV_M_COL] = np.nan
        mask2 = (df['time'] >= '12:07:00') & (df['time'] <= '12:22:00') & (df[NONINV_M_COL] > 80)
        df.loc[mask2, NONINV_M_COL] = np.nan
        mask3 = (df['time'] >= '09:22:00') & (df['time'] <= '09:37:00') & (df[NONINV_M_COL] > 100)
        df.loc[mask3, NONINV_M_COL] = np.nan
        mask4 = (df['time'] >= '13:07:00') & (df['time'] <= '13:10:00') & (df[NONINV_M_COL] > 100)
        df.loc[mask4, NONINV_M_COL] = np.nan
    
    return df


def filter_invasive_line(df_invasive, patient_name):
    """Create gaps in invasive line by removing specific time ranges"""
    if df_invasive is None or len(df_invasive) == 0:
        return df_invasive
    
    df = df_invasive.copy()
    
    if "Patient 49" in patient_name or "Patient_49" in patient_name:
        mask = (df['time'] >= '09:40:00') & (df['time'] <= '09:50:00')
        df.loc[mask, INV_M_COL] = np.nan
    
    if "PROMISES 51" in patient_name:
        mask = (df['time'] >= '09:15:00') & (df['time'] <= '09:25:00')
        df.loc[mask, INV_M_COL] = np.nan
    
    return df


def calculate_gradients(df_invasive, df_noninv):
    """Calculate gradient for each non-invasive point"""
    if df_invasive is None or df_noninv is None or len(df_invasive) == 0 or len(df_noninv) == 0:
        return None, None
    
    df_invasive['minutes'] = df_invasive['time'].apply(time_to_minutes)
    df_noninv['minutes'] = df_noninv['time'].apply(time_to_minutes)
    
    df_invasive = df_invasive.dropna(subset=['minutes', INV_M_COL])
    df_noninv = df_noninv.dropna(subset=['minutes', NONINV_M_COL])
    
    if len(df_invasive) == 0 or len(df_noninv) == 0:
        return None, None
    
    gradients = []
    times = []
    
    for _, pni_row in df_noninv.iterrows():
        t_pni = pni_row['minutes']
        val_pni = pni_row[NONINV_M_COL]
        
        idx = (df_invasive['minutes'] - t_pni).abs().idxmin()
        t_inv = df_invasive.loc[idx, 'minutes']
        val_inv = df_invasive.loc[idx, INV_M_COL]
        
        if abs(t_inv - t_pni) <= TIME_TOLERANCE:
            gradient = abs(val_inv - val_pni)
            gradients.append(gradient)
            times.append(t_pni)
    
    if len(gradients) == 0:
        return None, None
    
    return np.array(gradients), np.array(times)


def calculate_time_in_gradient(df_invasive, df_noninv, threshold=10):
    """Calculate time spent with gradient >= threshold"""
    gradients, times = calculate_gradients(df_invasive, df_noninv)
    
    if gradients is None or len(gradients) < 2:
        return 0, 0, 0, 0
    
    total_time = times[-1] - times[0]
    if total_time <= 0:
        return 0, 0, 0, 0
    
    time_in_gradient = 0
    
    for i in range(len(gradients) - 1):
        t1 = times[i]
        t2 = times[i + 1]
        g1 = gradients[i]
        g2 = gradients[i + 1]
        dt = t2 - t1
        
        if g1 >= threshold and g2 >= threshold:
            time_in_gradient += dt
        elif g1 >= threshold and g2 < threshold:
            if g2 != g1:
                fraction = (threshold - g1) / (g2 - g1)
                time_in_gradient += fraction * dt
        elif g1 < threshold and g2 >= threshold:
            if g2 != g1:
                fraction = (threshold - g1) / (g2 - g1)
                time_in_gradient += (1 - fraction) * dt
    
    percentage = (time_in_gradient / total_time) * 100 if total_time > 0 else 0
    
    return time_in_gradient, total_time, percentage, len(gradients)


def export_to_excel(results, output_path):
    """Export results to formatted Excel file"""
    
    df = pd.DataFrame(results)
    
    # Calculate additional statistics for summary
    all_percentages = df['percentage_gradient'].values
    all_medians = df['median_gradient'].values
    all_means = df['mean_gradient'].values
    all_maxs = df['max_gradient'].values
    
    # Create Excel writer
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        # Sheet 1: Patient results
        df_sorted = df.sort_values('percentage_gradient', ascending=False)
        df_sorted.to_excel(writer, sheet_name='Patients', index=False)
        
        # Sheet 2: Summary statistics
        summary_data = {
            'Statistic': [
                'Number of patients',
                'Mean percentage in gradient (%)',
                'Median percentage in gradient (%)',
                'Standard deviation percentage (%)',
                'Minimum percentage (%)',
                'Maximum percentage (%)',
                '25th percentile percentage (%)',
                '75th percentile percentage (%)',
                '',
                'Mean gradient (mmHg)',
                'Median gradient (mmHg)',
                'Standard deviation gradient (mmHg)',
                'Minimum gradient (mmHg)',
                'Maximum gradient (mmHg)',
                '25th percentile gradient (mmHg)',
                '75th percentile gradient (mmHg)',
                '',
                'Total time analyzed (hours)',
                'Mean time per patient (hours)',
                'Median time per patient (hours)'
            ],
            'Value': [
                len(df),
                round(np.mean(all_percentages), 1),
                round(np.median(all_percentages), 1),
                round(np.std(all_percentages), 1),
                round(np.min(all_percentages), 1),
                round(np.max(all_percentages), 1),
                round(np.percentile(all_percentages, 25), 1),
                round(np.percentile(all_percentages, 75), 1),
                '',
                round(np.mean(all_means), 1),
                round(np.median(all_medians), 1),
                round(np.std(all_means), 1),
                round(np.min(all_means), 1),
                round(np.max(all_means), 1),
                round(np.percentile(all_means, 25), 1),
                round(np.percentile(all_means, 75), 1),
                '',
                round(df['total_time_min'].sum() / 60, 1),
                round(df['total_time_min'].mean() / 60, 1),
                round(df['total_time_min'].median() / 60, 1)
            ]
        }
        df_summary = pd.DataFrame(summary_data)
        df_summary.to_excel(writer, sheet_name='Summary', index=False)
        
        # Sheet 3: High risk patients (>20% time in gradient)
        high_risk = df[df['percentage_gradient'] > 20].sort_values('percentage_gradient', ascending=False)
        if len(high_risk) > 0:
            high_risk.to_excel(writer, sheet_name='High_Risk_Patients', index=False)
        
        # Sheet 4: Percentile distribution
        percentile_data = {
            'Percentile': [10, 20, 25, 30, 40, 50, 60, 70, 75, 80, 90, 95, 99],
            'Percentage_in_gradient': [
                round(np.percentile(all_percentages, 10), 1),
                round(np.percentile(all_percentages, 20), 1),
                round(np.percentile(all_percentages, 25), 1),
                round(np.percentile(all_percentages, 30), 1),
                round(np.percentile(all_percentages, 40), 1),
                round(np.percentile(all_percentages, 50), 1),
                round(np.percentile(all_percentages, 60), 1),
                round(np.percentile(all_percentages, 70), 1),
                round(np.percentile(all_percentages, 75), 1),
                round(np.percentile(all_percentages, 80), 1),
                round(np.percentile(all_percentages, 90), 1),
                round(np.percentile(all_percentages, 95), 1),
                round(np.percentile(all_percentages, 99), 1)
            ],
            'Gradient_median': [
                round(np.percentile(all_medians, 10), 1),
                round(np.percentile(all_medians, 20), 1),
                round(np.percentile(all_medians, 25), 1),
                round(np.percentile(all_medians, 30), 1),
                round(np.percentile(all_medians, 40), 1),
                round(np.percentile(all_medians, 50), 1),
                round(np.percentile(all_medians, 60), 1),
                round(np.percentile(all_medians, 70), 1),
                round(np.percentile(all_medians, 75), 1),
                round(np.percentile(all_medians, 80), 1),
                round(np.percentile(all_medians, 90), 1),
                round(np.percentile(all_medians, 95), 1),
                round(np.percentile(all_medians, 99), 1)
            ]
        }
        df_percentiles = pd.DataFrame(percentile_data)
        df_percentiles.to_excel(writer, sheet_name='Percentiles', index=False)
    
    # Apply formatting
    wb = load_workbook(output_path)
    
    # Format Patients sheet
    ws = wb['Patients']
    
    # Header formatting
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='40466e', end_color='40466e', fill_type='solid')
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    
    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[col_letter].width = adjusted_width
    
    # Color code percentage column (column D)
    percent_col = None
    for col in range(1, ws.max_column + 1):
        if ws.cell(row=1, column=col).value == 'percentage_gradient':
            percent_col = col
            break
    
    if percent_col:
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=percent_col)
            try:
                val = float(cell.value)
                if val > 30:
                    cell.fill = PatternFill(start_color='ffcccc', end_color='ffcccc', fill_type='solid')
                elif val > 15:
                    cell.fill = PatternFill(start_color='ffffcc', end_color='ffffcc', fill_type='solid')
            except:
                pass
    
    # Color code median gradient column (column E)
    median_col = None
    for col in range(1, ws.max_column + 1):
        if ws.cell(row=1, column=col).value == 'median_gradient':
            median_col = col
            break
    
    if median_col:
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=median_col)
            try:
                val = float(cell.value)
                if val > 15:
                    cell.fill = PatternFill(start_color='ffcccc', end_color='ffcccc', fill_type='solid')
                elif val > 10:
                    cell.fill = PatternFill(start_color='ffffcc', end_color='ffffcc', fill_type='solid')
            except:
                pass
    
    # Format Summary sheet
    ws_summary = wb['Summary']
    for col in range(1, ws_summary.max_column + 1):
        cell = ws_summary.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    ws_summary.column_dimensions['A'].width = 35
    ws_summary.column_dimensions['B'].width = 20
    
    # Format Percentiles sheet
    ws_perc = wb['Percentiles']
    for col in range(1, ws_perc.max_column + 1):
        cell = ws_perc.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    for col in ws_perc.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws_perc.column_dimensions[col_letter].width = max_length + 2
    
    wb.save(output_path)


def main():
    print("=" * 80)
    print("CALCUL DU TEMPS PASSÉ EN GRADIENT")
    print(f"Gradient = |MAP - NBP| >= {GRADIENT_THRESHOLD} mmHg")
    print("=" * 80)
    
    # Get all invasive files
    invasive_files = list(INVASIVE_INPUT_DIR.glob("*_invasive_cleaned.csv"))
    print(f"\nFound {len(invasive_files)} invasive cleaned files")
    
    if len(invasive_files) == 0:
        print("\nWARNING: No invasive files found!")
        return
    
    # Get all non-invasive files
    noninvasive_files = list(NONINVASIVE_INPUT_DIR.glob("*_noninvasive.csv"))
    print(f"Found {len(noninvasive_files)} non-invasive files")
    
    # Create mapping
    patient_map = {}
    for f in invasive_files:
        base_name = f.stem.replace("_invasive_cleaned", "")
        patient_map[base_name] = {'invasive': f, 'noninvasive': None}
    
    for f in noninvasive_files:
        base_name = f.stem.replace("_noninvasive", "")
        if base_name in patient_map:
            patient_map[base_name]['noninvasive'] = f
    
    print(f"\nTotal unique patients: {len(patient_map)}")
    print()
    
    results = []
    
    for base_name, files in patient_map.items():
        print(f"Processing: {base_name}...", end=' ', flush=True)
        
        try:
            invasive_file = files['invasive']
            noninvasive_file = files['noninvasive']
            
            if invasive_file is None or noninvasive_file is None:
                print(f"SKIP - missing files")
                continue
            
            df_invasive = pd.read_csv(invasive_file)
            df_noninv = pd.read_csv(noninvasive_file)
            
            df_noninv = filter_noninvasive_by_time_and_value(df_noninv, base_name)
            df_invasive = filter_invasive_line(df_invasive, base_name)
            
            time_in_grad, total_time, percentage, n_points = calculate_time_in_gradient(df_invasive, df_noninv, GRADIENT_THRESHOLD)
            
            gradients, _ = calculate_gradients(df_invasive, df_noninv)
            
            if gradients is not None and len(gradients) > 0:
                results.append({
                    'patient': base_name,
                    'percentage_gradient': round(percentage, 1),
                    'time_in_gradient_min': round(time_in_grad, 1),
                    'total_time_min': round(total_time, 1),
                    'median_gradient': round(np.median(gradients), 1),
                    'mean_gradient': round(np.mean(gradients), 1),
                    'std_gradient': round(np.std(gradients), 1),
                    'max_gradient': round(np.max(gradients), 1),
                    'min_gradient': round(np.min(gradients), 1),
                    'q25_gradient': round(np.percentile(gradients, 25), 1),
                    'q75_gradient': round(np.percentile(gradients, 75), 1)
                })
            
            print(f"OK - {percentage:.1f}%")
            
        except Exception as e:
            print(f"ERROR: {str(e)[:80]}")
    
    if len(results) > 0:
        excel_output = OUTPUT_DIR / "gradient_statistics.xlsx"
        export_to_excel(results, excel_output)
        
        print("\n" + "=" * 80)
        print("RÉSULTATS")
        print("=" * 80)
        
        df_results = pd.DataFrame(results).sort_values('percentage_gradient', ascending=False)
        
        print(f"\n{'Patient':<35} {'% gradient':>12} {'Temps (min)':>14} {'Médiane':>10} {'Moyenne':>10}")
        print("-" * 85)
        
        for _, row in df_results.iterrows():
            print(f"{row['patient']:<35} {row['percentage_gradient']:>11.1f}% {row['time_in_gradient_min']:>13.1f} {row['median_gradient']:>9.1f} {row['mean_gradient']:>9.1f}")
        
        print(f"\nExcel file saved: {excel_output}")
    
    print("\n" + "=" * 80)


if __name__ == "__main__":
    main()